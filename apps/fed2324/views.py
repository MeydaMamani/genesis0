from django.shortcuts import render, get_object_or_404
from django.views.generic import TemplateView, View
from django.core import serializers
from django.http import JsonResponse, HttpResponse, QueryDict

from datetime import date, datetime
from .models import Teen, PackChild
from apps.redes.models import Redes

from django.db.models import Case, When, IntegerField, FloatField, ExpressionWrapper, Q, F, Sum, Count, IntegerField, Avg, Value, DecimalField
from django.db.models.functions import Cast, Round

# library excel
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Color

import datetime
import json
import locale


class MicroRedView(View):
    def get(self, request, *args, **kwargs):
        data = serializers.serialize('json', Redes.objects.filter(parent=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        data = json.loads(data)
        listMcrd = []
        for mcr_red in data:
            listMcrd.append({'pk': mcr_red['pk'], 'code': mcr_red['fields']['code'], 'name': mcr_red['fields']['name']})

        return HttpResponse(json.dumps(listMcrd), content_type='application/json')


class DistrictView(View):
    def get(self, request, *args, **kwargs):
        dist = serializers.serialize('json', Redes.objects.filter(parent=request.GET['id']), indent=2, use_natural_foreign_keys=True)
        dist = json.loads(dist)
        listDist = []
        for dist in dist:
            listDist.append({'pk': dist['pk'], 'code': dist['fields']['code'], 'name': dist['fields']['name']})

        return HttpResponse(json.dumps(listDist), content_type='application/json')


class Suple4View(TemplateView):
    template_name = 'kids/suple4m.html'


class TeenView(TemplateView):
    template_name = 'teen/index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['redes'] = Redes.objects.filter(type='P')
        return context


class ListTeen(TemplateView):
    def post(self, request, *args, **kwargs):
        print(request.POST)
        if self.request.session['sytem']['redLevel'] == '1':
            json_data4 = []
            if request.POST['red'] == 'TODOS':
                dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
                dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month']).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month']).aggregate(total=Count('id'))['total']
                cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], num=1).aggregate(cumplen=Count('id'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            elif request.POST['microred'] == 'TODOS':
                codigo = request.POST['red'].split("-")
                dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov= codigo[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
                dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov= codigo[1]).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov= codigo[1]).aggregate(total=Count('id'))['total']
                cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov= codigo[1], num=1).aggregate(cumplen=Count('id'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            elif request.POST['distrito'] == 'TODOS':
                cod_red = request.POST['red'].split("-")
                codigo = request.POST['microred'].split("-")
                dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=codigo[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
                dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=codigo[1]).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=codigo[1]).aggregate(total=Count('id'))['total']
                cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=codigo[1], num=1).aggregate(cumplen=Count('id'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            elif request.POST['distrito'] != 'TODOS':
                cod_red = request.POST['red'].split("-")
                cod_mcred = request.POST['microred'].split("-")
                dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito'])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
                dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito']).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito']).aggregate(total=Count('id'))['total']
                cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito'], num=1).aggregate(cumplen=Count('id'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            json_data4.append(dataNom)
            json_data4.append(list(dataDist))
            json_data4.append(dataTotal)
            return HttpResponse(json.dumps(json_data4), content_type='application/json')

        elif self.request.session['sytem']['redLevel'] == '2': #AQUI
            print(request.POST)
            json_data4 = []
            if request.POST['microred'] == 'TODOS':
                print('AQUI TOYmicror 2')
                codigo = request.POST['red'].split("-")
                dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov= codigo[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
                dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov= codigo[1]).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov= codigo[1]).aggregate(total=Count('id'))['total']
                cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov= codigo[1], num=1).aggregate(cumplen=Count('id'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            elif request.POST['distrito'] == 'TODOS':
                print('AQUI TOY 3')
                cod_red = request.POST['red'].split("-")
                codigo = request.POST['microred'].split("-")
                dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=codigo[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
                dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=codigo[1]).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=codigo[1]).aggregate(total=Count('id'))['total']
                cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=codigo[1], num=1).aggregate(cumplen=Count('id'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            elif request.POST['distrito'] != 'TODOS':
                print('AQUI TOY 4')
                cod_red = request.POST['red'].split("-")
                cod_mcred = request.POST['microred'].split("-")
                dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito'])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
                dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito']).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito']).aggregate(total=Count('id'))['total']
                cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito'], num=1).aggregate(cumplen=Count('id'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            json_data4.append(dataNom)
            json_data4.append(list(dataDist))
            json_data4.append(dataTotal)
            return HttpResponse(json.dumps(json_data4), content_type='application/json')

        elif self.request.session['sytem']['redLevel'] == '3':
            print(request.POST)
            json_data4 = []
            cod_red =  Redes.objects.get(id=request.POST['red'])
            cod_mcred = request.POST['microred'].split("-")
            if request.POST['distrito'] == 'TODOS':
                print('AQUI TOY 3')
                dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red.code, cod_microred=cod_mcred[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
                dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red.code, cod_microred=cod_mcred[1]).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red.code, cod_microred=cod_mcred[1]).aggregate(total=Count('id'))['total']
                cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_prov=cod_red.code, cod_microred=cod_mcred[1], num=1).aggregate(cumplen=Count('id'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            elif request.POST['distrito'] != 'TODOS':
                print('AQUI TOY 4')
                dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito'])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
                dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito']).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
                total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito']).aggregate(total=Count('id'))['total']
                cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_microred=cod_mcred[1], cod_dist=request.POST['distrito'], num=1).aggregate(cumplen=Count('id'))['cumplen']
                dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            json_data4.append(dataNom)
            json_data4.append(list(dataDist))
            json_data4.append(dataTotal)
            return HttpResponse(json.dumps(json_data4), content_type='application/json')

        elif self.request.session['sytem']['redLevel'] == '4':
            json_data4 = []
            dataNom = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_dist=self.request.session['sytem']['redCode'])
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))
            dataDist = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_dist=self.request.session['sytem']['redCode']).values('provincia', 'distrito').annotate(denominador=Count('id'),
                        numerador=Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())),
                        avance=(ExpressionWrapper( Cast(Sum(Case(When(num=1, then=1), default=0, output_field=IntegerField())), FloatField()) /
                        Cast(Count('*'), FloatField()) * 100, output_field=FloatField()))).order_by('-avance', '-denominador', '-numerador')
            total = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_dist=self.request.session['sytem']['redCode']).aggregate(total=Count('id'))['total']
            cumplen = Teen.objects.filter(anio=request.POST['year'], mes=request.POST['month'], cod_dist=self.request.session['sytem']['redCode'], num=1).aggregate(cumplen=Count('id'))['cumplen']
            dataTotal = { 'total': total, 'cumple': cumplen, 'avance': round((cumplen/total)*100, 1) if total != 0 else 0 }

            json_data4.append(dataNom)
            json_data4.append(list(dataDist))
            json_data4.append(dataTotal)
            return HttpResponse(json.dumps(json_data4), content_type='application/json')


class ReportTeenExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        # print(request.GET['red'], request.GET['microred'], request.GET['dist'], request.GET['year'], request.GET['month'])
        locale.setlocale(locale.LC_TIME, 'es_ES')
        nameMonth = datetime.date(1900, int(request.GET['month']), 1).strftime('%B')

        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:H2", "medium", "57267C")
        set_border(self, ws, "A4:H4", "medium", "366092")
        set_border(self, ws, "A6:H6", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.add_image(img, 'H2')

        ws.merge_cells('A2:G2')
        ws.row_dimensions[2].height = 33

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['C'].width = 32
        ws.column_dimensions['D'].width = 32
        ws.column_dimensions['F'].width = 13
        ws['A2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
        ws['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['A2'] = 'DIRESA PASCO DEIT: SI-05: Adolescentes mujeres de 12 a 17 años de edad, con dosaje de hemoglobina, en IPRESS del primer nivel de atención de salud del Gobierno Regional - ' + nameMonth.upper() + ' ' + request.GET['year']

        ws.merge_cells('A4:H4')
        ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['A4'] = 'CODIFICACION: Dosaje de Hemoglobina: 85018, 85018.01'

        ws.merge_cells('A6:H6')
        ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A6'] = 'Fuente: BD HisMinsa con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A8'] = '#'
        ws['A8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['B8'] = 'Provincia'
        ws['B8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['C8'] = 'Distrito'
        ws['C8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['D8'] = 'Establecimiento'
        ws['D8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['E8'] = 'Documento'
        ws['E8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['F8'] = 'Fecha Atención'
        ws['F8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['G8'] = 'Dosaje Hb'
        ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['G8'].fill = PatternFill(start_color='F7C3FA', end_color='F7C3FA', fill_type='solid')

        ws['H8'] = 'Cumple'
        ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        if self.request.session['sytem']['redLevel'] == '1':
            if request.GET['red'] == 'TODOS':
                dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.GET['microred'] == 'TODOS':
                codigo = request.GET['red'].split("-")
                dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'], cod_prov= codigo[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.GET['dist'] == 'TODOS':
                cod_red = request.GET['red'].split("-")
                codigo = request.GET['microred'].split("-")
                dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'], cod_prov=cod_red[1], cod_microred=codigo[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.GET['dist'] != 'TODOS':
                cod_red = request.GET['red'].split("-")
                cod_mcred = request.GET['microred'].split("-")
                dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.GET['dist'])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        elif self.request.session['sytem']['redLevel'] == '2': #AQUI
            print(request.GET)
            if request.GET['microred'] == 'TODOS':
                codigo = request.GET['red'].split("-")
                dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'], cod_prov= codigo[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.GET['dist'] == 'TODOS':
                print('AQUI TOY 3')
                cod_red = request.GET['red'].split("-")
                codigo = request.GET['microred'].split("-")
                dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'], cod_prov=cod_red[1], cod_microred=codigo[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.GET['dist'] != 'TODOS':
                print('AQUI TOY 4')
                cod_red = request.GET['red'].split("-")
                cod_mcred = request.GET['microred'].split("-")
                dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'], cod_prov=cod_red[1], cod_microred=cod_mcred[1], cod_dist=request.GET['dist'])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        elif self.request.session['sytem']['redLevel'] == '3':
            print(request.GET)
            cod_red =  Redes.objects.get(id=request.GET['red'])
            cod_mcred = request.GET['microred'].split("-")
            if request.GET['dist'] == 'TODOS':
                print('AQUI TOY 3')
                dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'], cod_prov=cod_red.code, cod_microred=cod_mcred[1])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

            elif request.GET['dist'] != 'TODOS':
                print('AQUI TOY 4')
                dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'], cod_microred=cod_mcred[1], cod_dist=request.GET['dist'])
                dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        elif self.request.session['sytem']['redLevel'] == '4':
            dataNom = Teen.objects.filter(anio=request.GET['year'], mes=request.GET['month'], cod_dist=self.request.session['sytem']['redCode'])
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 9
        cant = len(dataNom)
        num=1
        if cant > 0:
            for adol in dataNom:
                ws.cell(row=cont, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=1).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=1).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=1).value = num

                ws.cell(row=cont, column=2).alignment = Alignment(horizontal="left")
                ws.cell(row=cont, column=2).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=2).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=2).value = adol['fields']['provincia']

                ws.cell(row=cont, column=3).alignment = Alignment(horizontal="left")
                ws.cell(row=cont, column=3).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=3).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=3).value = adol['fields']['distrito']

                ws.cell(row=cont, column=4).alignment = Alignment(horizontal="left")
                ws.cell(row=cont, column=4).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=4).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=4).value = adol['fields']['establecimiento']

                ws.cell(row=cont, column=5).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=5).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=5).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=5).value = adol['fields']['documento']

                ws.cell(row=cont, column=6).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=6).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=6).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=6).value = adol['fields']['fec_1erAte']

                ws.cell(row=cont, column=7).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=7).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=7).font = Font(name='Calibri', size=9)
                ws.cell(row=cont, column=7).value = adol['fields']['fec_hb']

                if adol['fields']['num'] == '1':
                    cumplen = '✔'
                    ws.cell(row=cont, column=8).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=8).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=8).border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
                ws.cell(row=cont, column=8).value = cumplen

                cont = cont+1
                num = num+1

        # sheet2 = wb.create_sheet('RESUMEN')
        # sheet2['A1'] = 'SUSCRIPCION'
        nombre_archivo = "DEIT_PASCO ADOLESCENTES MUJERES CON DOSAJE HB.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL ADOLESCENTES'
        wb.save(response)
        return response


class PackChildView(TemplateView):
    template_name = 'kids/paquete.html'


class ReportPackChildExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        request.POST['year'] = request.GET['year']
        request.POST['month'] = request.GET['month']

        locale.setlocale(locale.LC_TIME, 'es_ES')
        nameMonth = datetime.date(1900, int(request.POST['month']), 1).strftime('%B')

        wb = Workbook()
        ws = wb.active

        def set_border(self, ws, cell_range, types, colors):
            thin = Side(border_style=types, color=colors)
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(self, ws, "A2:BC2", "medium", "57267C")
        set_border(self, ws, "A4:BC4", "medium", "366092")
        set_border(self, ws, "A6:BC6", "medium", "D9D9D9")

        img = Image('static/img/logoPrint.png')
        ws.add_image(img, 'A2')

        ws.merge_cells('B2:BC2')
        ws.row_dimensions[2].height = 34
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 21
        ws.column_dimensions['H'].width = 6
        ws.column_dimensions['AS'].width = 6
        ws.column_dimensions['AT'].width = 6
        ws.column_dimensions['AU'].width = 6
        ws.column_dimensions['AV'].width = 6
        ws.column_dimensions['AW'].width = 6
        ws.column_dimensions['AX'].width = 6
        ws.column_dimensions['AY'].width = 6
        ws.column_dimensions['AZ'].width = 6
        ws.column_dimensions['BA'].width = 7
        ws.column_dimensions['BB'].width = 7
        ws.column_dimensions['BC'].width = 7
        ws['B2'].font = Font(name='Aptos Narrow', size=11, bold=True, color='57267C')
        ws['B2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws['B2'] = 'DIRESA PASCO DEIT: MC-02: Niñas y niños menores de 12 meses de edad, procedentes de los quintiles 1 y 2 de pobreza departamental que reciben el paquete integrado de servicios - ' + nameMonth.upper() + ' ' + request.POST['year'] + ' A LA FECHA'

        ws.merge_cells('A4:BC4')
        ws['A4'].font = Font(name='Aptos Narrow', size=9, bold=True, color='305496')
        ws['A4'] = 'CODIFICACION: Cred Rn: 99381.01 - Cred Mes: 99381 - Vacuna Antipolio: 90712 - Vacuna Pentavalente: 90722 - Dx Anemia: D500, D508, D509, D649, D539 - Suplementación: 99199.17, 99199.19 - Prematuros: P073, P071, P0711, P00712 - Dosaje Hemoglobina: 85018, 85018.01'

        ws.merge_cells('A6:BC6')
        ws['A6'].font = Font(name='Aptos Narrow', size=9, bold=True, color='757171')
        ws['A6'] = 'Fuente: BD HisMinsa con Fecha: ' + date.today().strftime('%Y-%m-%d') + ' a las 08:30 horas'

        ws['A8'] = '#'
        ws['A8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['A8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['A8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['A8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['B8'] = 'Provincia'
        ws['B8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['B8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['B8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['C8'] = 'Distrito'
        ws['C8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['C8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['C8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['D8'] = 'Ult EESS Aten (His)'
        ws['D8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['D8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['D8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['D8'].fill = PatternFill(start_color='C3CBFA', end_color='C3CBFA', fill_type='solid')

        ws['E8'] = 'Documento'
        ws['E8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['E8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['E8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['F8'] = 'Apellidos y Nombres'
        ws['F8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['F8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['F8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['F8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['G8'] = 'Fecha Nacido'
        ws['G8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['G8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['G8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['G8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['H8'] = 'Seguro'
        ws['H8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['H8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['H8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['H8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['I8'] = '1° Ctrl'
        ws['I8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['I8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['I8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['I8'].fill = PatternFill(start_color='F7C3FA', end_color='F7C3FA', fill_type='solid')

        ws['J8'] = '2° Ctrl'
        ws['J8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['J8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['J8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['J8'].fill = PatternFill(start_color='F7C3FA', end_color='F7C3FA', fill_type='solid')

        ws['K8'] = '3° Ctrl'
        ws['K8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['K8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['K8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['K8'].fill = PatternFill(start_color='F7C3FA', end_color='F7C3FA', fill_type='solid')

        ws['L8'] = '4° Ctrl'
        ws['L8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['L8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['L8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['L8'].fill = PatternFill(start_color='F7C3FA', end_color='F7C3FA', fill_type='solid')

        ws['M8'] = 'Ctrls Rn'
        ws['M8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['M8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['M8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['M8'].fill = PatternFill(start_color='B3F5C2', end_color='B3F5C2', fill_type='solid')

        ws['N8'] = 'Cred 1'
        ws['N8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['N8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['N8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['N8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['O8'] = 'Cred 2'
        ws['O8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['O8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['O8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['O8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['P8'] = 'Antineumo 2M'
        ws['P8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['P8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['P8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['P8'].fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        ws['Q8'] = 'Rotavirus 2M'
        ws['Q8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Q8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Q8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Q8'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        ws['R8'] = 'Antipolio 2M'
        ws['R8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['R8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['R8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['R8'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')

        ws['S8'] = 'Penta 2M'
        ws['S8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['S8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['S8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['S8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['T8'] = 'Cred 3'
        ws['T8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['T8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['T8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['T8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['U8'] = 'Cred 4'
        ws['U8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['U8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['U8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['U8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['V8'] = 'Suple 4'
        ws['V8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['V8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['V8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['V8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['W8'] = 'Antineumo 4M'
        ws['W8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['W8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['W8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['W8'].fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

        ws['X8'] = 'Rotavirus 4M'
        ws['X8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['X8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['X8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['X8'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        ws['Y8'] = 'Penta 4M'
        ws['Y8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Y8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Y8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['Y8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['Z8'] = 'Antipolio 4M'
        ws['Z8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['Z8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['Z8'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws['Z8'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')

        ws['AA8'] = 'Cred 5'
        ws['AA8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AA8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AA8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AA8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AB8'] = 'Suple 5'
        ws['AB8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AB8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AB8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AB8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AC8'] = 'Cred 6'
        ws['AC8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AC8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AC8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AC8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AD8'] = 'Tamizaje'
        ws['AD8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AD8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AD8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AD8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AE8'] = 'Dx Anemia'
        ws['AE8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AE8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AE8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AE8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AF8'] = 'Suple 6'
        ws['AF8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AF8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AF8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AF8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AG8'] = 'Antipolio 6M'
        ws['AG8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AG8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AG8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AG8'].fill = PatternFill(start_color='F5F3CB', end_color='F5F3CB', fill_type='solid')

        ws['AH8'] = 'Penta 6M'
        ws['AH8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AH8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AH8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AH8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AI8'] = 'Cred 7'
        ws['AI8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AI8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AI8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AI8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AJ8'] = 'Suple 7'
        ws['AJ8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AJ8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AJ8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AJ8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AK8'] = 'Cred 8'
        ws['AK8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AK8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AK8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AK8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AL8'] = 'Suple 8'
        ws['AL8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AL8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AL8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AL8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AM8'] = 'Cred 9'
        ws['AM8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AM8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AM8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AM8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AN8'] = 'Suple 9'
        ws['AN8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AN8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AN8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AN8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AO8'] = 'Cred 10'
        ws['AO8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AO8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AO8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AO8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AP8'] = 'Suple 10'
        ws['AP8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AP8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AP8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AP8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AQ8'] = 'Cred 11'
        ws['AQ8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AQ8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AQ8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AQ8'].fill = PatternFill(start_color='DBCAF4', end_color='DBCAF4', fill_type='solid')

        ws['AR8'] = 'Suple 11'
        ws['AR8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AR8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AR8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AR8'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        ws['AS8'] = 'C1-C2'
        ws['AS8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AS8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AS8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AS8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AT8'] = 'C2-C3'
        ws['AT8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AT8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AT8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AT8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AU8'] = 'C3-C4'
        ws['AU8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AU8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AU8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AU8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AV8'] = 'C4-C5'
        ws['AV8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AV8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AV8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AV8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AW8'] = 'C5-C6'
        ws['AW8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AW8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AW8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AW8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AX8'] = 'C6-C7'
        ws['AX8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AX8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AX8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AX8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AY8'] = 'C7-C8'
        ws['AY8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AY8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AY8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AY8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['AZ8'] = 'C8-C9'
        ws['AZ8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['AZ8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['AZ8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['AZ8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['BA8'] = 'C9-C10'
        ws['BA8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BA8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BA8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BA8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['BB8'] = 'C10-C11'
        ws['BB8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BB8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BB8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BB8'].fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')

        ws['BC8'] = 'Cumple'
        ws['BC8'].border = Border(left=Side(border_style="thin", color="808080"), right=Side(border_style="thin", color="808080"), top=Side(border_style="thin", color="808080"), bottom=Side(border_style="thin", color="808080"))
        ws['BC8'].font = Font(name='Aptos Narrow', size=10, bold=True)
        ws['BC8'].alignment = Alignment(horizontal="center", vertical="center")
        ws['BC8'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        if self.request.session['sytem']['redLevel'] == '1':
            dataNom = PackChild.objects.filter(fec_nac__gte=datetime.date(int(request.POST['year']), int(request.POST['month']), 1))
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        elif self.request.session['sytem']['redLevel'] == '2': #AQUI
            dataNom = PackChild.objects.filter(fec_nac__gte=datetime.date(int(request.POST['year']), int(request.POST['month']), 1), provincia=self.request.session['sytem']['redName'])
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        elif self.request.session['sytem']['redLevel'] == '3':
            dataNom = PackChild.objects.filter(fec_nac__gte=datetime.date(int(request.POST['year']), int(request.POST['month']), 1), distrito=self.request.session['sytem']['redName'])
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        elif self.request.session['sytem']['redLevel'] == '4':
            dataNom = PackChild.objects.filter(fec_nac__gte=datetime.date(int(request.POST['year']), int(request.POST['month']), 1), distrito=self.request.session['sytem']['redName'])
            dataNom = json.loads(serializers.serialize('json', dataNom, indent=2, use_natural_foreign_keys=True))

        cont = 9
        cant = len(dataNom)
        num=1
        if cant > 0:
            for paqNinio in dataNom:
                ws.cell(row=cont, column=1).value = num
                ws.cell(row=cont, column=2).value = paqNinio['fields']['provincia']
                ws.cell(row=cont, column=3).value = paqNinio['fields']['distrito']
                ws.cell(row=cont, column=4).value = paqNinio['fields']['establecimiento']
                ws.cell(row=cont, column=5).value = paqNinio['fields']['documento']
                ws.cell(row=cont, column=6).value = paqNinio['fields']['ape_nombres']
                ws.cell(row=cont, column=7).value = paqNinio['fields']['fec_nac']
                ws.cell(row=cont, column=8).value = paqNinio['fields']['seguro']
                ws.cell(row=cont, column=8).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=9).value = paqNinio['fields']['ctrl1rn']
                ws.cell(row=cont, column=10).value = paqNinio['fields']['ctrl2rn']
                ws.cell(row=cont, column=11).value = paqNinio['fields']['ctrl3rn']
                ws.cell(row=cont, column=12).value = paqNinio['fields']['ctrl4rn']

                if paqNinio['fields']['cumplern'] == '1':
                    cumplen = '✔'
                    ws.cell(row=cont, column=13).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=13).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=13).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=13).value = cumplen
                ws.cell(row=cont, column=14).value = paqNinio['fields']['cred1']
                ws.cell(row=cont, column=15).value = paqNinio['fields']['cred2']
                ws.cell(row=cont, column=16).value = paqNinio['fields']['neumo2']
                ws.cell(row=cont, column=17).value = paqNinio['fields']['rota2']
                ws.cell(row=cont, column=18).value = paqNinio['fields']['polio2']
                ws.cell(row=cont, column=19).value = paqNinio['fields']['penta2']
                ws.cell(row=cont, column=20).value = paqNinio['fields']['cred3']
                ws.cell(row=cont, column=21).value = paqNinio['fields']['cred4']
                ws.cell(row=cont, column=22).value = paqNinio['fields']['suple4']
                ws.cell(row=cont, column=23).value = paqNinio['fields']['neumo4']
                ws.cell(row=cont, column=24).value = paqNinio['fields']['rota4']
                ws.cell(row=cont, column=25).value = paqNinio['fields']['penta4']
                ws.cell(row=cont, column=26).value = paqNinio['fields']['polio4']
                ws.cell(row=cont, column=27).value = paqNinio['fields']['cred5']
                ws.cell(row=cont, column=28).value = paqNinio['fields']['suple5']
                ws.cell(row=cont, column=29).value = paqNinio['fields']['cred6']
                ws.cell(row=cont, column=30).value = paqNinio['fields']['tmz']
                ws.cell(row=cont, column=31).value = paqNinio['fields']['dxAnemia']
                ws.cell(row=cont, column=32).value = paqNinio['fields']['suple6']
                ws.cell(row=cont, column=33).value = paqNinio['fields']['polio6']
                ws.cell(row=cont, column=34).value = paqNinio['fields']['penta6']
                ws.cell(row=cont, column=35).value = paqNinio['fields']['cred7']
                ws.cell(row=cont, column=36).value = paqNinio['fields']['suple7']
                ws.cell(row=cont, column=37).value = paqNinio['fields']['cred8']
                ws.cell(row=cont, column=38).value = paqNinio['fields']['suple8']
                ws.cell(row=cont, column=39).value = paqNinio['fields']['cred9']
                ws.cell(row=cont, column=40).value = paqNinio['fields']['suple9']
                ws.cell(row=cont, column=41).value = paqNinio['fields']['cred10']
                ws.cell(row=cont, column=42).value = paqNinio['fields']['suple10']
                ws.cell(row=cont, column=43).value = paqNinio['fields']['cred11']
                ws.cell(row=cont, column=44).value = paqNinio['fields']['suple11']
                ws.cell(row=cont, column=45).value = paqNinio['fields']['dif1']
                ws.cell(row=cont, column=46).value = paqNinio['fields']['dif2']
                ws.cell(row=cont, column=47).value = paqNinio['fields']['dif3']
                ws.cell(row=cont, column=48).value = paqNinio['fields']['dif4']
                ws.cell(row=cont, column=49).value = paqNinio['fields']['dif5']
                ws.cell(row=cont, column=50).value = paqNinio['fields']['dif6']
                ws.cell(row=cont, column=51).value = paqNinio['fields']['dif7']
                ws.cell(row=cont, column=52).value = paqNinio['fields']['dif8']
                ws.cell(row=cont, column=53).value = paqNinio['fields']['dif9']
                ws.cell(row=cont, column=54).value = paqNinio['fields']['dif10']

                if paqNinio['fields']['mide'] == '1':
                    cumplen = '✔'
                    ws.cell(row=cont, column=55).font = Font(name='Calibri', size=10, color='00AC4E')
                else:
                    cumplen = '✘'
                    ws.cell(row=cont, column=55).font = Font(name='Calibri', size=10, color='C00000')

                ws.cell(row=cont, column=55).alignment = Alignment(horizontal="center")
                ws.cell(row=cont, column=55).value = cumplen

                cont = cont+1
                num = num+1

        # sheet2 = wb.create_sheet('RESUMEN')
        # sheet2['A1'] = 'SUSCRIPCION'
        nombre_archivo = "DEIT_PASCO AVANCE PAQUETE NIÑO COMPLETO.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.title = 'NOMINAL PAQUETE NIÑO'
        wb.save(response)
        return response

